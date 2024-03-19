from openpyxl import Workbook
from decimal import Decimal
from openpyxl.styles import Font, Alignment
from typing import List
# from api_core import Item

def create_order_excel(order_list, order_id: int) -> str:
    wb = Workbook()
    sheet = wb.active
    sheet.title = 'Order Details'

    headers = ['Товар', 'Ссылка', 'Количество', 'Цена за единицу', 'Характеристики']
    header_font = Font(bold=True)
    alignment = Alignment(horizontal='center')

    row_num = 1
    for col_num, header in enumerate(headers, start=1):
        cell = sheet.cell(row=row_num, column=col_num, value=header)
        cell.font = header_font
        cell.alignment = alignment

    for item in order_list:
        row_num += 1
        row = [
            item['Товар'], 
            item['Ссылка'], 
            item['Количество'], 
            float(item['Цена за единицу']),  
            item['Характеристики']
        ]
        for col_num, value in enumerate(row, start=1):
            sheet.cell(row=row_num, column=col_num, value=value)
    
    filename = f"order_{order_id}_details.xlsx"
    
    try:
        wb.save(filename)
    except Exception as e:
        print(f"Error saving file: {e}")
        return ""

    return filename


def as_text(value):
    """Переводит значение в текст для вычисления ширины столбца."""
    if value is None:
        return ""
    return str(value)


# order_list = [
#     {'Товар': 'Laptop', 'Ссылка': 'http://testserver/shop/products/detail/laptop/', 'Количество': 1, 'Цена за единицу': Decimal('1000.00'), 'Характеристики': 'Характеристики продукта не указаны.'}
# ]

# create_order_excel(order_list, 1)
